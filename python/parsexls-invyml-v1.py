import openpyxl
import yaml
import ipaddress

def parse_excel_without_pandas(excel_file: str, output_file: str = 'inventory.yml') -> None:
    """
    Версия без использования pandas
    """
    try:
        # Загрузка Excel файла
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active  # Получаем активный лист
        
        if sheet is None:
            raise ValueError("Не удалось получить активный лист из Excel файла")
        
        inventory = {'all': {'hosts': {}}}
        
        # Получаем заголовки из первой строки
        headers = []
        for cell in sheet[1]:  # Первая строка (заголовки)
            headers.append(cell.value)
        
        # Проверяем необходимые столбцы
        required_columns = ['Name', 'HostName', 'Ip-address']
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Отсутствует столбец: {col}")
        
        # Индексы столбцов
        name_idx = headers.index('Name')
        hostname_idx = headers.index('HostName')
        ip_idx = headers.index('Ip-address')
        
        # Обработка строк данных (начиная со второй строки)
        row_count = 0
        for row_num in range(2, sheet.max_row + 1):
            try:
                # Получаем значения ячеек строки
                name_cell = sheet.cell(row=row_num, column=name_idx + 1).value
                if not name_cell:  # Пустое имя - пропускаем
                    continue
                
                host_name = str(name_cell).strip()
                hostname_cell = sheet.cell(row=row_num, column=hostname_idx + 1).value
                hostname = str(hostname_cell).strip() if hostname_cell else ""
                
                ip_cell = sheet.cell(row=row_num, column=ip_idx + 1).value
                ip_address = str(ip_cell).strip() if ip_cell else ""
                
                if not ip_address:
                    print(f"Пропущена строка {row_num}: отсутствует IP-адрес")
                    continue
                
                # Обработка IP
                try:
                    if '/' in ip_address:
                        ip_without_mask = ip_address.split('/')[0]
                    else:
                        ip_without_mask = ip_address
                    
                    # Проверяем валидность IP
                    ipaddress.ip_address(ip_without_mask)
                    
                except ValueError:
                    print(f"Пропущена строка {row_num}: некорректный IP-адрес '{ip_address}'")
                    continue
                
                # Создание записи
                host_vars = {
                    'ansible_host': ip_without_mask,
                    'ansible_user': 'admin'
                }
                
                if hostname:
                    host_vars['hostname'] = hostname
                
                inventory['all']['hosts'][host_name] = host_vars
                row_count += 1
                
            except Exception as e:
                print(f"Ошибка в строке {row_num}: {e}")
                continue
        
        # Запись в YAML
        with open(output_file, 'w', encoding='utf-8') as f:
            yaml.dump(inventory, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
        
        print(f"✅ Inventory файл успешно создан: {output_file}")
        print(f"📊 Обработано хостов: {row_count}")
        
    except FileNotFoundError:
        print(f"❌ Ошибка: Файл '{excel_file}' не найден")
    except Exception as e:
        print(f"❌ Ошибка: {e}")

# Альтернативная версия с явным указанием листа
def parse_excel_explicit_sheet(excel_file: str, output_file: str = 'inventory.yml', sheet_name: str = None) -> None:
    """
    Версия с явным указанием листа
    """
    try:
        # Загрузка Excel файла
        workbook = openpyxl.load_workbook(excel_file)
        
        # Получаем лист
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        if sheet is None:
            available_sheets = workbook.sheetnames
            raise ValueError(f"Активный лист не найден. Доступные листы: {', '.join(available_sheets)}")
        
        print(f"📄 Обрабатываем лист: {sheet.title}")
        
        inventory = {'all': {'hosts': {}}}
        
        # Получаем заголовки
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Column_{col}")
        
        print(f"📋 Заголовки: {headers}")
        
        # Проверяем необходимые столбцы
        required_columns = ['Name', 'HostName', 'Ip-address']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise ValueError(f"Отсутствуют столбцы: {', '.join(missing_columns)}")
        
        # Индексы столбцов
        name_idx = headers.index('Name')
        hostname_idx = headers.index('HostName')
        ip_idx = headers.index('Ip-address')
        
        # Обработка строк
        row_count = 0
        for row_num in range(2, sheet.max_row + 1):
            try:
                name_val = sheet.cell(row=row_num, column=name_idx + 1).value
                if not name_val:
                    continue
                
                host_name = str(name_val).strip()
                hostname_val = sheet.cell(row=row_num, column=hostname_idx + 1).value
                hostname = str(hostname_val).strip() if hostname_val else ""
                
                ip_val = sheet.cell(row=row_num, column=ip_idx + 1).value
                ip_address = str(ip_val).strip() if ip_val else ""
                
                if not ip_address:
                    print(f"⚠️ Пропущена строка {row_num}: отсутствует IP-адрес")
                    continue
                
                # Обработка IP
                try:
                    if '/' in ip_address:
                        ip_without_mask = ip_address.split('/')[0].strip()
                    else:
                        ip_without_mask = ip_address.strip()
                    
                    ipaddress.ip_address(ip_without_mask)
                    
                except ValueError as e:
                    print(f"⚠️ Пропущена строка {row_num}: некорректный IP-адрес '{ip_address}'")
                    continue
                
                # Создание записи хоста
                host_vars = {
                    'ansible_host': ip_without_mask,
                    'ansible_user': 'admin'  # Замените на нужного пользователя
                }
                
                if hostname:
                    host_vars['hostname'] = hostname
                
                inventory['all']['hosts'][host_name] = host_vars
                row_count += 1
                print(f"✅ Добавлен хост: {host_name} -> {ip_without_mask}")
                
            except Exception as e:
                print(f"❌ Ошибка в строке {row_num}: {e}")
                continue
        
        # Проверяем, что есть хосты
        if not inventory['all']['hosts']:
            print("⚠️ Внимание: не найдено ни одного валидного хоста")
        
        # Запись в YAML
        with open(output_file, 'w', encoding='utf-8') as f:
            yaml.dump(inventory, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
        
        print(f"✅ Inventory файл успешно создан: {output_file}")
        print(f"📊 Обработано хостов: {row_count}")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")

# Создание тестового Excel файла
def create_sample_excel():
    """Создает пример Excel файла для тестирования"""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Hosts"
        
        # Заголовки
        sheet['A1'] = 'Name'
        sheet['B1'] = 'HostName'
        sheet['C1'] = 'Ip-address'
        
        # Данные
        data = [
            ['web-server-01', 'web01.local', '192.168.1.10/24'],
            ['web-server-02', 'web02.local', '192.168.1.11/24'],
            ['db-server-01', 'db01.local', '192.168.1.20/24'],
        ]
        
        for i, row_data in enumerate(data, start=2):
            sheet[f'A{i}'] = row_data[0]
            sheet[f'B{i}'] = row_data[1]
            sheet[f'C{i}'] = row_data[2]
        
        workbook.save('hosts.xlsx')
        print("✅ Создан пример файла: hosts.xlsx")
        print("📋 Структура файла:")
        print("   - Лист: 'Hosts'")
        print("   - Столбцы: Name, HostName, Ip-address")
        
    except Exception as e:
        print(f"❌ Ошибка при создании файла: {e}")

# Запуск
if __name__ == "__main__":
    # Создаем тестовый файл (раскомментируйте если нужно)
    create_sample_excel()
    
    # Запускаем парсер
    print("🚀 Запуск парсера...")
    parse_excel_explicit_sheet('hosts.xlsx', 'inventory.yml')
    
    # Или простую версию
    # parse_excel_without_pandas('hosts.xlsx', 'inventory.yml')