#!/usr/bin/env python3
"""
Упрощенный парсер Excel для hosts.yml
"""

import sys
from openpyxl import load_workbook
import yaml

def main():
    if len(sys.argv) < 2:
        print("Использование: python3 script.py <excel-file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Загружаем Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Получаем заголовки
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1) if ws.cell(1, col).value]
    
    print("Столбцы:", ", ".join(headers))
    column = input("Введите имя столбца с хостами: ").strip()
    
    if column not in headers:
        print("Ошибка: Столбец не найден")
        sys.exit(1)
    
    # Собираем хосты
    col_idx = headers.index(column) + 1
    hosts = []
    
    for row in range(2, ws.max_row + 1):
        host = ws.cell(row, col_idx).value
        if host:
            hosts.append(str(host).strip())
    
    # Сохраняем
    with open('hosts.yml', 'w') as f:
        yaml.dump({'hosts': ','.join(hosts)}, f)
    
    print(f"Создан hosts.yml с {len(hosts)} хостами")

if __name__ == "__main__":
    main()